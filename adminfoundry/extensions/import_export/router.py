"""Export endpoint — CSV always, XLSX when openpyxl is installed.

``GET /api/v1/admin/{resource}/_export?format=csv``
``GET /api/v1/admin/{resource}/_export?format=xlsx``

Query parameters:
    ``format``  required, ``csv`` or ``xlsx``
    ``ids``     optional, repeated. If given, exports ONLY those primary
                keys (and ignores ``search``). Used by the UI to export
                the user's row selection.
    ``search``  optional, same semantics as the list endpoint. Ignored
                when ``ids`` is present.
    ``limit``   optional, max rows, capped at :data:`MAX_EXPORT_ROWS`.

Authorization: identical to ``admin.<resource>.list``. Anyone who can
list a resource can also export it.

The ``xlsx`` format requires ``openpyxl`` — install via
``pip install adminfoundry[xlsx]``. If openpyxl is missing the endpoint
returns 501 with a helpful message.
"""

from __future__ import annotations

import csv
import io
import logging
from typing import Any

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Query,
    Request,
    Response,
    UploadFile,
    status,
)
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from adminfoundry.admin.context import AdminContext, require_admin_context
from adminfoundry.audit import record_audit_in_session, request_audit_kwargs
from adminfoundry.authz.permissions import permission_key
from adminfoundry.crud.payload import clean_write_payload
from adminfoundry.crud.query import (
    apply_filters,
    apply_ordering,
    apply_search,
    coerce_primary_key_value,
    parse_filter_query,
    primary_key_column,
)
from adminfoundry.db.dependencies import get_async_session
from adminfoundry.registry import ModelAdmin
from adminfoundry.schemas.builder import build_model_schema
from adminfoundry.schemas.serialization.serializer import serialize_records
from adminfoundry.security.validation import (
    InvalidResourceNameError,
    validate_resource_name,
)

logger = logging.getLogger(__name__)

#: Hard cap on rows per export — prevents accidental DoS via the public
#: endpoint. Configurable per request via ``?limit=`` up to this value.
MAX_EXPORT_ROWS: int = 10_000

#: Audit action string written for every export.
EXPORT_AUDIT_ACTION: str = "crud_export"

#: Audit action string written for every import.
IMPORT_AUDIT_ACTION: str = "crud_import"

#: Supported export formats; xlsx is only usable when openpyxl is installed.
SUPPORTED_EXPORT_FORMATS: tuple[str, ...] = ("csv", "xlsx")

#: Hard cap on imported rows per request. Each row triggers a flush, so
#: keep this materially smaller than MAX_EXPORT_ROWS.
MAX_IMPORT_ROWS: int = 5_000


def xlsx_available() -> bool:
    """True when ``openpyxl`` is importable (so the xlsx format works).

    Checked without importing openpyxl itself — the dependency stays
    lazy, loaded only inside the export/import handlers.
    """
    import importlib.util

    return importlib.util.find_spec("openpyxl") is not None


def available_formats() -> tuple[str, ...]:
    """Formats this install can actually export/import.

    CSV is always available; XLSX needs ``openpyxl``. The extension
    publishes this through the contract so the UI never offers a format
    the server would answer with 501.
    """
    return ("csv", "xlsx") if xlsx_available() else ("csv",)


def _resolve_admin(request: Request, resource: str) -> type[ModelAdmin]:
    try:
        resource = validate_resource_name(resource)
    except InvalidResourceNameError:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Resource '{resource}' is not registered.",
        ) from None
    admin = request.app.state.adminfoundry.registry.get(resource)
    if admin is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Resource '{resource}' is not registered.",
        )
    return admin


def _require_list_permission(ctx: AdminContext, resource: str) -> None:
    if ctx.tenant is None:
        return  # superadmin / no tenant context — handled by the request-time gate
    required = permission_key(resource, "list")
    if not ctx.has_permission(required):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=f"Missing required permission: {required}",
        )


def _columns_for(admin: ModelAdmin, sample_row: dict[str, Any] | None) -> list[str]:
    """Choose the CSV header. ``list_display`` wins if non-empty; otherwise
    fall back to the keys of the first serialized row, which already
    excludes protected/hidden fields."""
    if admin.list_display:
        return list(admin.list_display)
    if sample_row is not None:
        return list(sample_row.keys())
    return []


def _build_csv(rows: list[dict[str, Any]], columns: list[str]) -> str:
    buffer = io.StringIO()
    writer = csv.DictWriter(
        buffer,
        fieldnames=columns,
        extrasaction="ignore",
        quoting=csv.QUOTE_MINIMAL,
    )
    writer.writeheader()
    for row in rows:
        # csv.DictWriter writes empty for missing keys; coerce non-stringables.
        writer.writerow({c: _cell(row.get(c)) for c in columns})
    return buffer.getvalue()


def _build_xlsx(rows: list[dict[str, Any]], columns: list[str]) -> bytes:
    """Lazy openpyxl import — keeps the extension importable without it."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - exercised via the 501 path
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail=(
                "XLSX export requires openpyxl. Install with "
                "'pip install adminfoundry[xlsx]' and restart the app."
            ),
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    for row in rows:
        ws.append([_xlsx_cell(row.get(c)) for c in columns])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return str(value)
    return str(value)


def _xlsx_cell(value: Any) -> Any:
    """Pass primitives through unchanged so openpyxl preserves types
    (numbers stay numbers in Excel); stringify everything else."""
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


router = APIRouter()


@router.get("/{resource}/_export")
async def export_records(
    resource: str,
    request: Request,
    format: str = "csv",
    search: str | None = None,
    limit: int = MAX_EXPORT_ROWS,
    ids: list[str] = Query(default_factory=list),
    session: AsyncSession = Depends(get_async_session),
    ctx: AdminContext = Depends(require_admin_context),
) -> Response:
    if format not in SUPPORTED_EXPORT_FORMATS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Unsupported export format: {format!r}. "
                f"Supported: {', '.join(SUPPORTED_EXPORT_FORMATS)}."
            ),
        )

    admin = _resolve_admin(request, resource)
    _require_list_permission(ctx, admin.model_name)

    capped_limit = max(1, min(int(limit), MAX_EXPORT_ROWS))

    base_stmt = select(admin.model)
    if ids:
        # Selection-based export: ignore search and the limit (the limit is
        # still applied as a safety cap, but with the typical UI selection
        # sizes this is a no-op).
        if len(ids) > MAX_EXPORT_ROWS:
            raise HTTPException(
                status_code=413,
                detail=(
                    f"Too many ids: {len(ids)} (cap: {MAX_EXPORT_ROWS})."
                ),
            )
        pk_col = primary_key_column(admin.model)
        try:
            coerced = [coerce_primary_key_value(admin.model, str(raw)) for raw in ids]
        except HTTPException:
            # coerce_primary_key_value raises 422; on export an invalid id is
            # a client mistake on the URL, surface as 400.
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="One or more 'ids' values are not valid for this resource.",
            ) from None
        base_stmt = base_stmt.where(pk_col.in_(coerced))
        base_stmt = apply_ordering(base_stmt, admin).limit(capped_limit)
    else:
        # D4: respect the same ``?filter_<field>=value`` query params
        # the list endpoint accepts so "export current view" is a
        # 1:1 wire equivalent of the on-screen filtered list.
        filters = parse_filter_query(request.query_params, admin)
        base_stmt = apply_filters(base_stmt, admin, filters)
        base_stmt = apply_search(base_stmt, admin, search)
        base_stmt = apply_ordering(base_stmt, admin).limit(capped_limit)

    records = (await session.execute(base_stmt)).scalars().all()
    rows = serialize_records(records, admin)
    columns = _columns_for(admin, rows[0] if rows else None)

    if format == "csv":
        body: bytes | str = _build_csv(rows, columns)
        media_type = "text/csv; charset=utf-8"
        filename = f"{admin.model_name}.csv"
    else:  # xlsx — may raise 501 if openpyxl missing
        body = _build_xlsx(rows, columns)
        media_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"{admin.model_name}.xlsx"

    try:
        await record_audit_in_session(
            session,
            action=EXPORT_AUDIT_ACTION,
            actor=ctx.principal,
            resource=admin.model_name,
            tenant_id=ctx.tenant.id if ctx.tenant is not None else None,
            changes={
                "rows": len(rows),
                "format": format,
                "search": search,
                "selected_ids": len(ids) if ids else 0,
                "filters": {
                    k: v
                    for k, v in (
                        request.query_params.multi_items()
                        if hasattr(request.query_params, "multi_items")
                        else list(request.query_params)
                    )
                    if k.startswith("filter_")
                },
            },
            **request_audit_kwargs(request, status_code=200),
        )
    except Exception:
        logger.warning(
            "export audit hook failed for resource=%s format=%s",
            admin.model_name, format, exc_info=True,
        )

    return Response(
        content=body,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------


def _require_create_permission(ctx: AdminContext, resource: str) -> None:
    if ctx.tenant is None:
        return
    required = permission_key(resource, "create")
    if not ctx.has_permission(required):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=f"Missing required permission: {required}",
        )


def _detect_import_format(filename: str) -> str:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return "csv"
    if name.endswith(".xlsx"):
        return "xlsx"
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=(
            "Unsupported import file extension. Upload a .csv or .xlsx file."
        ),
    )


def _parse_csv_upload(raw: bytes) -> list[dict[str, Any]]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx_upload(raw: bytes) -> list[dict[str, Any]]:
    """Lazy openpyxl import — same 501 path as the export side."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised via the 501 path
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail=(
                "XLSX import requires openpyxl. Install with "
                "'pip install adminfoundry[xlsx]' and restart the app."
            ),
        ) from exc

    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return []
    headers: list[str] = [str(h) if h is not None else "" for h in header]

    result: list[dict[str, Any]] = []
    for row in rows_iter:
        item: dict[str, Any] = {}
        for idx, name in enumerate(headers):
            if not name:
                continue
            item[name] = row[idx] if idx < len(row) else None
        if any(v is not None and v != "" for v in item.values()):
            result.append(item)
    return result


def _normalize_import_row(row: dict[str, Any]) -> dict[str, Any]:
    """Drop empty-string cells so they fall through to model defaults / NULL
    instead of being submitted as the literal empty string. CSV has no way to
    distinguish "" from NULL; we treat empty as missing, which is what users
    expect when round-tripping an exported file."""
    return {k: v for k, v in row.items() if not (isinstance(v, str) and v == "")}


def _format_error_detail(detail: Any) -> str:
    if isinstance(detail, str):
        return detail
    if isinstance(detail, dict):
        msg = detail.get("message", "Validation error")
        fields = detail.get("fields")
        if fields:
            return f"{msg} ({', '.join(map(str, fields))})"
        return str(msg)
    return str(detail)


@router.post("/{resource}/_import")
async def import_records(
    resource: str,
    request: Request,
    file: UploadFile = File(...),
    dry_run: bool = Query(
        False,
        description=(
            "Validate + simulate every row without persisting anything "
            "(Roadmap 5.3). Same response shape; the ``dry_run`` flag "
            "in the body tells the caller nothing was committed."
        ),
    ),
    session: AsyncSession = Depends(get_async_session),
    ctx: AdminContext = Depends(require_admin_context),
) -> dict[str, Any]:
    admin = _resolve_admin(request, resource)
    _require_create_permission(ctx, admin.model_name)

    filename = file.filename or "upload"
    fmt = _detect_import_format(filename)
    raw = await file.read()

    if fmt == "csv":
        rows = _parse_csv_upload(raw)
    else:
        rows = _parse_xlsx_upload(raw)

    if len(rows) > MAX_IMPORT_ROWS:
        # 413 — note: Starlette aliases CONTENT_TOO_LARGE to the old
        # REQUEST_ENTITY_TOO_LARGE; we pass the int directly to avoid the
        # deprecation warning either name triggers on the in-use version.
        raise HTTPException(
            status_code=413,
            detail=(
                f"Too many rows: {len(rows)} (cap: {MAX_IMPORT_ROWS}). "
                "Split the file or raise the limit."
            ),
        )

    schema = build_model_schema(admin)
    created = 0
    errors: list[dict[str, Any]] = []

    # Use a context labelled ``source="import"`` so lifecycle hooks can
    # branch on origin (e.g. skip an outbound webhook when rows are
    # being bulk-imported). ``dataclasses.replace`` is a shallow copy —
    # principal/tenant/permissions are shared with the original ctx by
    # reference, which is fine since they are frozen / read-only.
    from dataclasses import replace as _replace

    import_ctx = _replace(ctx, source="import")

    # Roadmap 5.3: wrap the dry-run loop in a savepoint we explicitly
    # roll back at the end. This is the only sound way to discard the
    # flushed rows while leaving the outer request transaction healthy
    # — calling ``session.rollback()`` directly would tear down the
    # outer transaction and the request-end commit would re-flush any
    # state still in the identity map.
    batch_sp = await session.begin_nested() if dry_run else None

    try:
        for idx, raw_row in enumerate(rows, start=1):
            try:
                async with session.begin_nested():
                    row_payload: dict[str, Any] = _normalize_import_row(raw_row)
                    row_payload = await admin.before_validate(row_payload, import_ctx)
                    cleaned = clean_write_payload(row_payload, schema, partial=False)
                    await admin.validate_create(cleaned, import_ctx)
                    cleaned = await admin.before_create(cleaned, import_ctx)
                    record = admin.model(**cleaned)
                    session.add(record)
                    await session.flush()
                    await admin.after_create(record, import_ctx)
                created += 1
            except HTTPException as exc:
                errors.append({"row": idx, "error": _format_error_detail(exc.detail)})
            except Exception as exc:
                errors.append({"row": idx, "error": str(exc) or exc.__class__.__name__})
    finally:
        if batch_sp is not None and batch_sp.is_active:
            await batch_sp.rollback()

    failed = len(errors)

    if dry_run:
        # Audit is intentionally skipped — a dry-run isn't a real
        # action and operators don't want every "let me check this
        # file" probe in the audit log.
        return {
            "dry_run": True,
            "created": created,
            "failed": failed,
            "total": created + failed,
            "errors": errors,
        }

    try:
        await record_audit_in_session(
            session,
            action=IMPORT_AUDIT_ACTION,
            actor=ctx.principal,
            resource=admin.model_name,
            tenant_id=ctx.tenant.id if ctx.tenant is not None else None,
            changes={
                "created": created,
                "failed": failed,
                "format": fmt,
                "filename": filename,
            },
            **request_audit_kwargs(request, status_code=200),
        )
    except Exception:
        logger.warning(
            "import audit hook failed for resource=%s format=%s",
            admin.model_name, fmt, exc_info=True,
        )

    return {
        "dry_run": False,
        "created": created,
        "failed": failed,
        "total": created + failed,
        "errors": errors,
    }


